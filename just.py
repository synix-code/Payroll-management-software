"""
ui.py
=====
Main PyQt6 user interface for Synix — Attendance & Payroll Manager.
Cosmo Hydraulic Industries.

Features
--------
- Upload .dat files (processed in a background worker thread, req #23)
- Employee list with search + filter
- Per-employee attendance detail view (full month calendar-style table)
- Salary breakdown (per_day, per_hour, OT, Sunday bonus, deductions, net)
- Admin panel:
    - Edit employee info (name, salary, advance)
    - Mark holidays for any date
    - Approve suspicious shifts
    - Manual override of any attendance day (req #24)
    - Add manual punch (IN/OUT correction)
- Signature setup for PDF salary slips
- PDF salary slip generation (4×4 inch)
- Audit log viewer
- Sidebar summary statistics

Architecture  (req #23 — responsive UI)
----------------------------------------
File parsing runs in a QThread worker so the UI never freezes.
All attendance + payroll logic is in attendance_engine.py / payroll_engine.py.
DataStore (validators.py) handles all persistence with atomic save.
"""

import sys
import os
import logging
import urllib.request
import json as _json
from datetime import datetime
from typing import Dict, List, Optional

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QFrame, QDialog, QLineEdit, QFormLayout, QMessageBox,
    QCheckBox, QDoubleSpinBox, QComboBox, QSizePolicy, QScrollArea,
    QGroupBox, QTextEdit, QTabWidget, QSplitter, QSpinBox, QProgressBar,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QColor, QPalette, QFont

from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import inch
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from attendance_engine import (
    process_dat_file, mark_holiday, approve_suspicious_shift,
    manual_override, add_manual_punch, run_engine_tests,
    DayStatus, ShiftState, get_audit_log, clear_audit_log,
)
from payroll_engine import calc_salary, calc_all_salaries, run_payroll_tests
from validators import DataStore, find_suspicious_shifts, find_missing_salary_employees

logger = logging.getLogger("synix.ui")

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
APP_NAME        = "Synix"
COMPANY_NAME    = "Cosmo Hydraulic Industries"
DEVELOPER_NAME  = "Synix Lab"
DEVELOPER_PHONE = "+91 7408096938"

# ─────────────────────────────────────────────────────────────────────────────
#  LICENSE / KILL-SWITCH  — checked at startup via GitHub JSON
# ─────────────────────────────────────────────────────────────────────────────
LICENSE_URL = (
    "https://raw.githubusercontent.com/synix-code/Synix-payroll-management/main/status.json"
)
LICENSE_TIMEOUT = 6   # seconds


def fetch_license() -> dict:
    """
    Fetch the license JSON from GitHub.

    Returns
    -------
    {"ok": True,  "enabled": True/False, "message": str}   on success
    {"ok": False, "reason": "no_internet" | "error", "detail": str}  on failure
    """
    try:
        req = urllib.request.Request(
            LICENSE_URL,
            headers={"User-Agent": f"{APP_NAME}/1.0"},
        )
        with urllib.request.urlopen(req, timeout=LICENSE_TIMEOUT) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
            return {
                "ok":      True,
                "enabled": bool(data.get("enabled", True)),
                "message": str(data.get("message", "")),
                
            }
        
    except urllib.error.URLError as e:
        # No internet / DNS failure / connection refused
        reason = "no_internet"
        return {"ok": False, "reason": reason, "detail": str(e)}
    except Exception as e:
        return {"ok": False, "reason": "error", "detail": str(e)}


class LicenseCheckDialog(QDialog):
    """
    Shown at startup while the license JSON is being fetched in a background
    thread, and again if the result is a block or a network error.

    States
    ------
    checking  : spinner + "Verifying…"
    disabled  : red icon + message from JSON  → Exit button only
    no_internet: orange icon + retry / exit buttons
    """

    _result_ready = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(APP_NAME)
        self.setFixedSize(420, 280)
        self.setWindowFlags(
            Qt.WindowType.Dialog | Qt.WindowType.CustomizeWindowHint |
            Qt.WindowType.WindowTitleHint
        )
        self.setStyleSheet(STYLE)

        self._lo = QVBoxLayout(self)
        self._lo.setContentsMargins(32, 28, 32, 28)
        self._lo.setSpacing(16)

        # Icon label (emoji-based, no external assets needed)
        self._icon = QLabel("⏳")
        self._icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._icon.setStyleSheet("font-size: 36px; background: transparent;")
        self._lo.addWidget(self._icon)

        self._title = QLabel("Verifying license…")
        self._title.setObjectName("heading")
        self._title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lo.addWidget(self._title)

        self._msg = QLabel("")
        self._msg.setWordWrap(True)
        self._msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._msg.setStyleSheet(f"color:{COLORS['muted']}; font-size:12px; background:transparent;")
        self._lo.addWidget(self._msg)

        self._btn_row = QHBoxLayout()
        self._lo.addLayout(self._btn_row)

        self._result_ready.connect(self._on_result)
        self._start_check()

    # ── background fetch ──────────────────────────────────────────────────────
    def _start_check(self):
        self._set_state("checking")
        self._worker = _LicenseWorker()
        self._worker.done.connect(self._result_ready)
        self._worker.start()

    def _set_state(self, state: str, message: str = ""):
        # clear buttons
        while self._btn_row.count():
            item = self._btn_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if state == "checking":
            self._icon.setText("⏳")
            self._title.setText("Verifying license…")
            self._msg.setText("Please wait, connecting to server.")

        elif state == "disabled":
            self._icon.setText("🚫")
            self._title.setText("Access Blocked")
            self._title.setStyleSheet(
                f"font-size:20px; font-weight:700; color:{COLORS['danger']}; background:transparent;"
            )
            self._msg.setText(message or "This application has been disabled by the administrator.")
            exit_btn = QPushButton("Exit")
            exit_btn.setObjectName("danger")
            exit_btn.clicked.connect(lambda: sys.exit(0))
            self._btn_row.addStretch()
            self._btn_row.addWidget(exit_btn)
            self._btn_row.addStretch()

        elif state == "no_internet":
            self._icon.setText("🌐")
            self._title.setText("No Internet Connection")
            self._title.setStyleSheet(
                f"font-size:18px; font-weight:700; color:{COLORS['warning']}; background:transparent;"
            )
            self._msg.setText(
                "Could not connect to the license server.\n"
                "Please connect to the internet and retry."
            )
            retry_btn = QPushButton("🔄  Retry")
            retry_btn.clicked.connect(self._start_check)
            exit_btn = QPushButton("Exit")
            exit_btn.setObjectName("danger")
            exit_btn.clicked.connect(lambda: sys.exit(0))
            self._btn_row.addStretch()
            self._btn_row.addWidget(retry_btn)
            self._btn_row.addWidget(exit_btn)
            self._btn_row.addStretch()

    # ── result handler ────────────────────────────────────────────────────────
    def _on_result(self, result: dict):
        if not result["ok"]:
            # Network error or no internet
            self._set_state("no_internet")
            return

        if not result["enabled"]:
            # Kill-switch active
            self._set_state("disabled", result.get("message", ""))
            return

        # All good — accept and let main window open
        self.accept()


class _LicenseWorker(QThread):
    """Fetch license JSON in background so UI stays responsive."""
    done = pyqtSignal(dict)

    def run(self):
        result = fetch_license()
        self.done.emit(result)

COLORS = {
    "bg":       "#0F1117",
    "surface":  "#1A1D2E",
    "card":     "#252840",
    "accent":   "#6C63FF",
    "success":  "#4CAF82",
    "warning":  "#FFB347",
    "danger":   "#FF5C5C",
    "text":     "#E8E9F3",
    "muted":    "#9395A5",
    "border":   "#2E3152",
    "sunday":   "#3D2E52",
    "saturday": "#2E3D52",
    "absent":   "#3D2E2E",
    "present":  "#2E3D34",
    "holiday":  "#1E3A2E",
}

STATUS_BG_FG = {
    DayStatus.PRESENT.value:         (COLORS["present"],  COLORS["success"]),
    DayStatus.HALF_DAY.value:        (COLORS["present"],  COLORS["warning"]),
    DayStatus.SATURDAY.value:        (COLORS["saturday"], COLORS["warning"]),
    DayStatus.SUNDAY.value:          (COLORS["sunday"],   "#BB86FC"),
    DayStatus.SUNDAY_ISOLATED.value: ("#2E2040",          "#F59E0B"),
    DayStatus.SUNDAY_OFF.value:      (COLORS["bg"],       COLORS["muted"]),
    DayStatus.SATURDAY_OFF.value:    (COLORS["bg"],       COLORS["muted"]),
    DayStatus.ABSENT.value:          (COLORS["absent"],   COLORS["danger"]),
    DayStatus.SUNDAY_ABSENT.value:   ("#3D2828",          COLORS["danger"]),
    DayStatus.HOLIDAY.value:         (COLORS["holiday"],  COLORS["success"]),
    DayStatus.ERROR.value:           (COLORS["absent"],   COLORS["danger"]),
}
STATUS_LABEL = {
    DayStatus.PRESENT.value:         "Present",
    DayStatus.HALF_DAY.value:        "Half Day",
    DayStatus.SATURDAY.value:        "Saturday",
    DayStatus.SUNDAY.value:          "Sunday Work",
    DayStatus.SUNDAY_ISOLATED.value: "Sunday OT-Only",
    DayStatus.SUNDAY_OFF.value:      "Sunday Off",
    DayStatus.SATURDAY_OFF.value:    "Saturday Off",
    DayStatus.ABSENT.value:          "Absent",
    DayStatus.SUNDAY_ABSENT.value:   "Sunday (Absent)",
    DayStatus.HOLIDAY.value:         "Holiday",
    DayStatus.ERROR.value:           "Data Error",
}

STYLE = f"""
QMainWindow, QDialog {{
    background: {COLORS['bg']};
}}
QWidget {{
    background: {COLORS['bg']};
    color: {COLORS['text']};
    font-family: 'Segoe UI', 'SF Pro Display', sans-serif;
    font-size: 13px;
}}
QFrame#card {{
    background: {COLORS['card']};
    border-radius: 12px;
    border: 1px solid {COLORS['border']};
}}
QFrame#surface {{
    background: {COLORS['surface']};
    border-radius: 8px;
    border: 1px solid {COLORS['border']};
}}
QPushButton {{
    background: {COLORS['accent']};
    color: white;
    border: none;
    border-radius: 8px;
    padding: 8px 18px;
    font-weight: 600;
    font-size: 13px;
}}
QPushButton:hover {{ background: #7B73FF; }}
QPushButton:pressed {{ background: #5A52E0; }}
QPushButton#danger {{ background: {COLORS['danger']}; }}
QPushButton#danger:hover {{ background: #FF7070; }}
QPushButton#success {{ background: {COLORS['success']}; }}
QPushButton#success:hover {{ background: #5DCF93; }}
QPushButton#secondary {{
    background: {COLORS['surface']};
    border: 1px solid {COLORS['border']};
    color: {COLORS['text']};
}}
QPushButton#secondary:hover {{ background: {COLORS['card']}; }}
QPushButton#warning {{
    background: {COLORS['warning']};
    color: #1A1A1A;
}}
QLineEdit, QDoubleSpinBox, QSpinBox, QComboBox, QTextEdit {{
    background: {COLORS['surface']};
    border: 1px solid {COLORS['border']};
    border-radius: 8px;
    padding: 8px 12px;
    color: {COLORS['text']};
    font-size: 13px;
}}
QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus, QComboBox:focus {{
    border: 1px solid {COLORS['accent']};
}}
QTableWidget {{
    background: {COLORS['surface']};
    border: 1px solid {COLORS['border']};
    border-radius: 10px;
    gridline-color: {COLORS['border']};
    color: {COLORS['text']};
}}
QTableWidget::item {{
    padding: 8px;
    border-bottom: 1px solid {COLORS['border']};
}}
QTableWidget::item:selected {{
    background: rgba(108,99,255,0.3);
    color: white;
}}
QHeaderView::section {{
    background: {COLORS['card']};
    color: {COLORS['muted']};
    padding: 10px 8px;
    border: none;
    border-bottom: 2px solid {COLORS['accent']};
    font-weight: 600;
    font-size: 12px;
}}
QScrollBar:vertical {{
    background: {COLORS['surface']};
    width: 6px;
    border-radius: 3px;
}}
QScrollBar::handle:vertical {{
    background: {COLORS['border']};
    border-radius: 3px;
    min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{ background: {COLORS['accent']}; }}
QScrollBar:horizontal {{
    background: {COLORS['surface']};
    height: 6px;
    border-radius: 3px;
}}
QScrollBar::handle:horizontal {{
    background: {COLORS['border']};
    border-radius: 3px;
}}
QTabWidget::pane {{
    border: 1px solid {COLORS['border']};
    border-radius: 10px;
    background: {COLORS['surface']};
}}
QTabBar::tab {{
    background: {COLORS['card']};
    color: {COLORS['muted']};
    padding: 10px 20px;
    border-radius: 8px 8px 0 0;
    margin-right: 2px;
    font-weight: 500;
}}
QTabBar::tab:selected {{
    background: {COLORS['accent']};
    color: white;
}}
QGroupBox {{
    border: 1px solid {COLORS['border']};
    border-radius: 10px;
    margin-top: 16px;
    padding-top: 8px;
    color: {COLORS['muted']};
    font-weight: 600;
    font-size: 12px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 12px;
    top: -8px;
    background: {COLORS['bg']};
    padding: 0 6px;
    color: {COLORS['accent']};
}}
QCheckBox {{
    color: {COLORS['text']};
    spacing: 8px;
}}
QCheckBox::indicator {{
    width: 18px; height: 18px;
    border-radius: 4px;
    border: 2px solid {COLORS['border']};
    background: {COLORS['surface']};
}}
QCheckBox::indicator:checked {{
    background: {COLORS['accent']};
    border-color: {COLORS['accent']};
}}
QLabel#heading {{
    font-size: 22px; font-weight: 700; color: {COLORS['text']};
}}
QLabel#subheading {{
    font-size: 15px; font-weight: 600; color: {COLORS['muted']};
}}
QLabel#stat_value {{
    font-size: 26px; font-weight: 700; color: {COLORS['text']};
}}
QLabel#stat_label {{
    font-size: 11px; font-weight: 600; color: {COLORS['muted']};
    text-transform: uppercase;
}}
QProgressBar {{
    background: {COLORS['surface']};
    border: 1px solid {COLORS['border']};
    border-radius: 6px;
    height: 8px;
    text-align: center;
}}
QProgressBar::chunk {{
    background: {COLORS['accent']};
    border-radius: 5px;
}}
"""

# ─────────────────────────────────────────────────────────────────────────────
#  GLOBAL DATA STORE INSTANCE
# ─────────────────────────────────────────────────────────────────────────────
store = DataStore()


# ─────────────────────────────────────────────────────────────────────────────
#  BACKGROUND WORKER THREAD  (req #23 — responsive UI)
# ─────────────────────────────────────────────────────────────────────────────
class DatParserWorker(QThread):
    """
    Parses a .dat file in a background thread so the UI stays responsive.
    Emits finished(attendance, month_label) on success or error(message) on failure.
    """
    finished = pyqtSignal(dict, str)
    error    = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, filepath: str, existing_holidays: dict) -> None:
        super().__init__()
        self.filepath          = filepath
        self.existing_holidays = existing_holidays

    def run(self) -> None:
        try:
            self.progress.emit("Parsing .dat file...")
            att, ml = process_dat_file(self.filepath, self.existing_holidays)
            self.progress.emit("Processing complete.")
            self.finished.emit(att, ml)
        except Exception as e:
            logger.exception("DatParserWorker error")
            self.error.emit(str(e))


# ─────────────────────────────────────────────────────────────────────────────
#  REUSABLE WIDGETS
# ─────────────────────────────────────────────────────────────────────────────
class StatCard(QFrame):
    """Small stat card widget: label on top, big value below, coloured."""

    def __init__(self, label: str, value, color: str = None, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(4)

        lbl = QLabel(label.upper())
        lbl.setObjectName("stat_label")
        lbl.setAlignment(Qt.AlignmentFlag.AlignLeft)

        self.val_lbl = QLabel(str(value))
        self.val_lbl.setObjectName("stat_value")
        self.val_lbl.setAlignment(Qt.AlignmentFlag.AlignLeft)
        if color:
            self.val_lbl.setStyleSheet(
                f"color: {color}; font-size: 26px; font-weight: 700; background: transparent;"
            )
        layout.addWidget(lbl)
        layout.addWidget(self.val_lbl)

    def update_value(self, val) -> None:
        self.val_lbl.setText(str(val))


# ─────────────────────────────────────────────────────────────────────────────
#  PAYROLL EXCEL REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
def generate_payroll_excel(output_path: str) -> str:
    """
    Generate an attractive Excel payroll report for all employees.

    Columns: SR | Emp ID | Name | Salary | Total Working Days |
             Total OT Count | Advanced Paid | Total Payable | Paid ✓

    Header: Cosmo Hydraulic Industries
    Footer row: Total salary summary
    """
    from payroll_engine import calc_all_salaries

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    # ── Colour palette ────────────────────────────────────────────────────────
    C_HEADER_BG   = "12153D"   # dark navy
    C_HEADER_FG   = "FFFFFF"
    C_ACCENT      = "6C63FF"   # purple accent
    C_COL_HDR_BG  = "1E2352"   # slightly lighter navy
    C_COL_HDR_FG  = "FFFFFF"
    C_ALT_ROW     = "F0F0FA"   # light lavender for alternating rows
    C_WHITE       = "FFFFFF"
    C_PAID_BG     = "DCFCE7"   # green tint
    C_PAID_FG     = "15803D"
    C_UNPAID_BG   = "FEE2E2"   # red tint
    C_UNPAID_FG   = "991B1B"
    C_TOTAL_BG    = "12153D"   # same as header
    C_TOTAL_FG    = "4ADE80"   # bright green
    C_BORDER      = "C7C7E0"

    thin  = Side(style="thin",   color=C_BORDER)
    thick = Side(style="medium", color=C_ACCENT)
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(left=thin, right=thin, top=thick, bottom=thick)

    def style_cell(cell, bold=False, fg=None, bg=None, size=11,
                   halign="center", valign="center", border=None, italic=False, wrap=False):
        cell.font      = Font(name="Arial", bold=bold, italic=italic,
                              color=fg or "000000", size=size)
        if bg:
            cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                   wrap_text=wrap)
        if border:
            cell.border = border

    # ── Row 1: Company header (merged across all 9 columns) ───────────────────
    month = store.month_label or datetime.now().strftime("%B %Y")
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 36

    ws.merge_cells("A1:I1")
    hdr = ws["A1"]
    hdr.value = "COSMO HYDRAULIC INDUSTRIES"
    style_cell(hdr, bold=True, fg=C_HEADER_FG, bg=C_HEADER_BG, size=18, halign="center")
    hdr.border = Border(left=thick, right=thick, top=thick, bottom=thin)

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"Payroll Report  —  {month}"
    style_cell(sub, bold=False, italic=True, fg="8B8FC8", bg=C_HEADER_BG, size=11, halign="center")
    sub.border = Border(left=thick, right=thick, top=thin, bottom=thick)

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 30
    COLS = [
        ("SR.",              8),
        ("Emp ID",          10),
        ("Name",            24),
        ("Salary (Rs)",     14),
        ("Working Days",    14),
        ("Total OT (hrs)",  14),
        ("Advance Paid",    14),
        ("Total Payable",   15),
        ("Paid ✓",          10),
    ]
    for col_idx, (header, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        style_cell(cell, bold=True, fg=C_COL_HDR_FG, bg=C_COL_HDR_BG,
                   size=10, halign="center", border=cell_border)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ─────────────────────────────────────────────────────────────
    eids      = store.all_eids()
    payrolls  = calc_all_salaries(store.employees, store.attendance, store.month_label)
    data_start = 4
    total_payable = 0.0

    for sr, eid in enumerate(eids, start=1):
        row  = data_start + sr - 1
        info = store.emp_info(eid)
        sc   = payrolls.get(eid, {})
        paid = info.get("paid", False)

        ws.row_dimensions[row].height = 22
        bg = C_ALT_ROW if sr % 2 == 0 else C_WHITE

        work_days  = (sc.get("work_days", 0) + sc.get("half_days", 0)
                      + sc.get("sun_days", 0) + sc.get("sun_iso_days", 0))
        total_ot   = sc.get("total_ot_h", 0)
        advance    = sc.get("advance", 0)
        net        = sc.get("net", 0)
        salary_cfg = sc.get("salary", info.get("salary", 0))

        total_payable += net

        row_data = [
            sr,
            eid,
            info.get("name", eid),
            salary_cfg,
            work_days,
            round(total_ot, 2),
            advance,
            net,
            "✅ Paid" if paid else "❌ Unpaid",
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            is_money = col_idx in (4, 7, 8)
            is_paid_col = col_idx == 9

            if is_paid_col:
                paid_bg = C_PAID_BG if paid else C_UNPAID_BG
                paid_fg = C_PAID_FG if paid else C_UNPAID_FG
                style_cell(cell, bold=True, fg=paid_fg, bg=paid_bg,
                           size=10, halign="center", border=cell_border)
            else:
                style_cell(cell, bold=(col_idx == 8), fg="1A1A3E", bg=bg,
                           size=10, halign="right" if is_money else "center",
                           border=cell_border)
                if is_money and val:
                    cell.number_format = '#,##0'

    # ── Total row ─────────────────────────────────────────────────────────────
    total_row = data_start + len(eids)
    ws.row_dimensions[total_row].height = 28

    ws.merge_cells(f"A{total_row}:G{total_row}")
    lbl = ws[f"A{total_row}"]
    lbl.value = f"TOTAL PAYABLE  ({len(eids)} Employees)"
    style_cell(lbl, bold=True, fg=C_TOTAL_FG, bg=C_TOTAL_BG,
               size=12, halign="center", border=bottom_border)

    total_cell = ws.cell(row=total_row, column=8, value=total_payable)
    style_cell(total_cell, bold=True, fg=C_TOTAL_FG, bg=C_TOTAL_BG,
               size=13, halign="right", border=bottom_border)
    total_cell.number_format = '#,##0'

    paid_count   = sum(1 for e in eids if store.emp_info(e).get("paid"))
    unpaid_count = len(eids) - paid_count
    summary_cell = ws.cell(row=total_row, column=9,
                           value=f"✅{paid_count} / ❌{unpaid_count}")
    style_cell(summary_cell, bold=True, fg=C_HEADER_FG, bg=C_TOTAL_BG,
               size=10, halign="center", border=bottom_border)

    # ── Freeze top rows ───────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
#  PDF SALARY SLIP GENERATOR  (4×4 inch)
# ─────────────────────────────────────────────────────────────────────────────
def generate_salary_slip(eid: str, output_path: str) -> str:
    """
    Generate a 4×4 inch PDF salary slip for an employee.
    Uses the global store for employee info and attendance.

    Returns
    -------
    output_path (same as input, for chaining)
    """
    info = store.emp_info(eid)
    att  = store.attendance.get(str(eid), {})
    sc   = calc_salary(eid, info, att, store.month_label)
    month = store.month_label or datetime.now().strftime("%B %Y")

    W = H = 4 * inch
    c = rl_canvas.Canvas(output_path, pagesize=(W, H))

    def hx(h):
        return colors.HexColor(h)

    # ── Section heights ───────────────────────────────────────────────────────
    HEADER_H = 0.76 * inch
    FOOTER_H = 0.18 * inch
    EMP_H    = 0.22 * inch
    CHIP_H   = 0.26 * inch
    ROW_H    = 0.17 * inch
    NET_H    = 0.36 * inch
    SIG_H    = 0.48 * inch
    PAD      = 0.10 * inch
    n_rows   = 4

    lx = 0.16 * inch
    rx = W - 0.16 * inch
    MX = 0.10 * inch

    # ── Layout: top-down so nothing ever overlaps ─────────────────────────────
    # ReportLab y=0 is bottom, y=H is top. We compute all positions top-down
    # then each section is a fixed band — zero chance of overlap.
    y_header_bot = H - HEADER_H

    y_emp_bot    = y_header_bot - PAD
    y_emp_top    = y_emp_bot - EMP_H          # not used directly but clear

    y_chip_bot   = y_emp_bot - EMP_H - PAD * 0.5
    y_chip_top   = y_chip_bot - CHIP_H

    y_table_top  = y_chip_bot - CHIP_H - PAD * 0.5
    y_table_bot  = y_table_top - n_rows * ROW_H
    y_divider    = y_table_top + PAD * 0.25   # thin line just above table

    y_net_bot    = y_table_bot - PAD - NET_H      # net box bottom: fully below table
    y_net_top    = y_net_bot + NET_H              # top of net pay box (RL y goes up)

    y_sig_bot    = y_net_bot - PAD                # signature starts below net box

    # Background
    c.setFillColor(hx('#FAFAFA'))
    c.rect(0, 0, W, H, fill=1, stroke=0)

    # ── Header ──────────────────────────────────────────────────────────────
    c.setFillColor(hx('#12153D'))
    c.rect(0, y_header_bot, W, HEADER_H, fill=1, stroke=0)
    # Decorative corner
    p = c.beginPath()
    p.moveTo(W - 0.9 * inch, H)
    p.lineTo(W, H)
    p.lineTo(W, y_header_bot + 0.3 * inch)
    p.close()
    c.setFillColor(hx('#1E2352'))
    c.drawPath(p, fill=1, stroke=0)
    # Accent line
    c.setFillColor(hx('#6C63FF'))
    c.rect(0, y_header_bot, W, 0.028 * inch, fill=1, stroke=0)
    # "SALARY SLIP" pill
    c.setFillColor(hx('#6C63FF'))
    c.roundRect(lx, y_header_bot + 0.44 * inch, 0.65 * inch, 0.16 * inch, 5, fill=1, stroke=0)
    c.setFont('Helvetica-Bold', 5.5)
    c.setFillColor(hx('#FFFFFF'))
    c.drawCentredString(lx + 0.325 * inch, y_header_bot + 0.49 * inch, 'SALARY SLIP')
    # Company name
    c.setFont('Helvetica-Bold', 10.5)
    c.setFillColor(hx('#FFFFFF'))
    c.drawCentredString(W / 2, y_header_bot + 0.30 * inch, 'COSMO HYDRAULIC INDUSTRIES')
    c.setFont('Helvetica', 7)
    c.setFillColor(hx('#8B8FC8'))
    c.drawCentredString(W / 2, y_header_bot + 0.12 * inch, f'{month.upper()}')

    # ── Employee info row ────────────────────────────────────────────────────
    emp_text_y = y_emp_bot - EMP_H * 0.45    # centred in emp band
    c.setFont('Helvetica-Bold', 8)
    c.setFillColor(hx('#1A1A3E'))
    c.drawString(lx, emp_text_y, info.get("name", eid).upper())
    c.setFont('Helvetica', 6.5)
    c.setFillColor(hx('#6C63FF'))
    c.drawRightString(rx, emp_text_y, f'EMP ID: {eid}')

    # Divider (sits just above the chip row)
    c.setStrokeColor(hx('#E2E2EE'))
    c.setLineWidth(0.6)
    c.line(lx, y_chip_bot, rx, y_chip_bot)

    # ── Attendance chips ─────────────────────────────────────────────────────
    chips = [
        (f"{sc['work_days']+sc['half_days']} Days",  '#22C55E', '#DCFCE7'),
        (f"{sc['absent_days']} Absent",              '#EF4444', '#FEE2E2'),
        (f"{sc['total_ot_h']:.1f}h OT",              '#F59E0B', '#FEF3C7'),
    ]
    if sc['sun_days'] > 0:
        chips.append((f"{sc['sun_days']} Sun", '#8B5CF6', '#EDE9FE'))

    n   = len(chips)
    cw  = min(0.72 * inch, (W - 2 * lx - (n - 1) * 0.05 * inch) / n)
    gap = 0.05 * inch
    cx0 = (W - n * cw - (n - 1) * gap) / 2
    chip_pill_h = 0.19 * inch
    cy  = y_chip_bot - CHIP_H + (CHIP_H - chip_pill_h) / 2   # vertically centred in chip band
    for i, (txt, fg, bg) in enumerate(chips):
        cx = cx0 + i * (cw + gap)
        c.setFillColor(hx(bg))
        c.roundRect(cx, cy, cw, chip_pill_h, 5, fill=1, stroke=0)
        c.setFont('Helvetica-Bold', 6)
        c.setFillColor(hx(fg))
        c.drawCentredString(cx + cw / 2, cy + chip_pill_h * 0.35, txt)

    # ── Earnings table ───────────────────────────────────────────────────────
    rows_data = [
        ("Basic Earned",              f"Rs {sc['weekday_earned']:,.0f}",      False),
        ("Sunday Bonus",              f"Rs {sc['sun_base_pay']+sc['sun_ot_pay']+sc['iso_ot_pay']:,.0f}", False),
        (f"OT Pay ({sc['total_ot_h']:.1f}h)", f"Rs {sc['weekday_ot_pay']:,.0f}", False),
        ("Advance Deduction",         f"-Rs {sc['advance']:,.0f}",            True),
    ]
    for i, (label, val, is_ded) in enumerate(rows_data):
        rb  = y_table_top - (i + 1) * ROW_H   # bottom of this row
        ty  = rb + ROW_H * 0.38               # text baseline — centred in row
        if i % 2 == 0:
            c.setFillColor(hx('#F0F0FA'))
            c.rect(MX, rb, W - 2 * MX, ROW_H, fill=1, stroke=0)
        c.setFont('Helvetica', 6.5)
        c.setFillColor(hx('#444466'))
        c.drawString(lx + 0.04 * inch, ty, label)
        c.setFont('Helvetica-Bold', 6.5)
        c.setFillColor(hx('#EF4444') if is_ded else hx('#1A1A3E'))
        c.drawRightString(rx - 0.04 * inch, ty, val)

    # ── Net pay box ──────────────────────────────────────────────────────────
    c.setFillColor(hx('#12153D'))
    c.roundRect(MX, y_net_bot, W - 2 * MX, NET_H, 6, fill=1, stroke=0)
    c.setFillColor(hx('#6C63FF'))
    c.roundRect(MX, y_net_bot, 0.04 * inch, NET_H, 3, fill=1, stroke=0)
    net_mid = y_net_bot + NET_H / 2
    c.setFont('Helvetica', 6.5)
    c.setFillColor(hx('#8B8FC8'))
    c.drawString(lx + 0.10 * inch, net_mid + 0.05 * inch, 'NET PAYABLE AMOUNT')
    c.setFont('Helvetica-Bold', 11)
    c.setFillColor(hx('#4ADE80'))
    c.drawRightString(rx - 0.04 * inch, net_mid - 0.07 * inch, f'Rs {sc["net"]:,.0f}')

    # ── Signature zone ───────────────────────────────────────────────────────
    sig_x1   = W - 1.1 * inch
    sig_x2   = W - 0.14 * inch
    sig_line = y_sig_bot - SIG_H + 0.30 * inch   # horizontal line position
    if store.sig_path and os.path.exists(store.sig_path):
        try:
            c.drawImage(
                ImageReader(store.sig_path),
                sig_x1, sig_line + 0.02 * inch,
                width=(sig_x2 - sig_x1), height=0.26 * inch,
                preserveAspectRatio=True, mask='auto',
            )
        except Exception:
            pass
    c.setStrokeColor(hx('#BBBBCC'))
    c.setLineWidth(0.6)
    c.line(sig_x1, sig_line, sig_x2, sig_line)
    c.setFont('Helvetica', 5.5)
    c.setFillColor(hx('#999999'))
    c.drawCentredString((sig_x1 + sig_x2) / 2, sig_line - 0.09 * inch, 'Authorised Signatory')

    # PAID stamp — only if salary is actually marked as paid
    if info.get("paid", False):
        c.saveState()
        c.translate(0.58 * inch, sig_line + 0.10 * inch)
        c.rotate(10)
        c.setStrokeColor(hx('#22C55E'))
        c.setFillColor(hx('#DCFCE7'))
        c.setLineWidth(1.5)
        c.roundRect(-0.30 * inch, -0.12 * inch, 0.60 * inch, 0.24 * inch, 4, fill=1, stroke=1)
        c.setFont('Helvetica-Bold', 8.5)
        c.setFillColor(hx('#15803D'))
        c.drawCentredString(0, -0.036 * inch, 'PAID')
        c.restoreState()

    # ── Footer ───────────────────────────────────────────────────────────────
    c.setFillColor(hx('#EBEBF5'))
    c.rect(0, 0, W, FOOTER_H, fill=1, stroke=0)
    c.setFont('Helvetica', 5)
    c.setFillColor(hx('#AAAACC'))
    c.drawCentredString(
        W / 2, FOOTER_H * 0.38,
        f'Powered by {DEVELOPER_NAME}  \u2022  {DEVELOPER_PHONE}',
    )

    c.save()
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
#  THERMAL SALARY SLIP  — TSC TE244 optimised  (4 × 3 inch, white bg, black ink)
# ─────────────────────────────────────────────────────────────────────────────
def generate_thermal_slip(eid: str, output_path: str) -> str:
    """
    Generate a thermal-printer-friendly salary slip for TSC TE244.

    Design rules for thermal printing
    ----------------------------------
    - Pure white background  (no wasted ink / ribbon)
    - Black text + thin lines only  (crisp at 203 DPI)
    - Zero filled rectangles with dark colour
    - No gradients, no tints, no decorative fills
    - Font sizes >= 6.5pt so 203 DPI renders cleanly
    - Page size: 4 inch wide x 3 inch tall (fits TE244 max width 4.25 in)

    Returns output_path.
    """
    info  = store.emp_info(eid)
    att   = store.attendance.get(str(eid), {})
    sc    = calc_salary(eid, info, att, store.month_label)
    month = store.month_label or datetime.now().strftime("%B %Y")

    W  = 4.00 * inch
    H  = 3.00 * inch
    c  = rl_canvas.Canvas(output_path, pagesize=(W, H))

    BLACK = colors.HexColor('#000000')
    WHITE = colors.HexColor('#FFFFFF')
    GRAY  = colors.HexColor('#555555')
    LX    = 0.14 * inch
    RX    = W - 0.14 * inch
    CX    = W / 2

    def txt(x, y, text, size=7, bold=False, align='left'):
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
        c.setFillColor(BLACK)
        if align == 'center':   c.drawCentredString(x, y, text)
        elif align == 'right':  c.drawRightString(x, y, text)
        else:                    c.drawString(x, y, text)

    def hline(y, x1=None, x2=None, lw=0.5, dashed=False):
        c.setStrokeColor(BLACK)
        c.setLineWidth(lw)
        c.setDash(2, 2) if dashed else c.setDash()
        c.line(x1 or LX, y, x2 or RX, y)

    def outline_box(x, y, w, h, lw=0.6):
        c.setStrokeColor(BLACK)
        c.setFillColor(WHITE)
        c.setLineWidth(lw)
        c.rect(x, y, w, h, fill=0, stroke=1)

    # White background
    c.setFillColor(WHITE)
    c.rect(0, 0, W, H, fill=1, stroke=0)

    # ── layout cursor (ReportLab y=0 bottom, y=H top) ────────────────────────
    y = H

    # 1. Header
    y -= 0.06 * inch
    txt(CX, y - 0.14 * inch, 'COSMO HYDRAULIC INDUSTRIES', size=9, bold=True, align='center')
    y -= 0.18 * inch
    txt(CX, y - 0.10 * inch, f'SALARY SLIP  |  {month.upper()}', size=6.5, align='center')
    y -= 0.14 * inch
    hline(y, lw=1.0)
    y -= 0.05 * inch

    # 2. Employee row
    txt(LX, y - 0.11 * inch, info.get('name', eid).upper(), size=8, bold=True)
    txt(RX, y - 0.11 * inch, f'EMP ID: {eid}', size=7, align='right')
    y -= 0.16 * inch
    salary_cfg = sc.get('salary', 0)
    txt(LX, y - 0.10 * inch, f'Monthly Salary: Rs {salary_cfg:,.0f}', size=6.5)
    txt(RX, y - 0.10 * inch, f'Per Day: Rs {sc["per_day"]:,.0f}  Per Hr: Rs {sc["per_hour"]:.2f}', size=6.5, align='right')
    y -= 0.14 * inch
    hline(y, dashed=True)
    y -= 0.06 * inch

    # 3. Attendance summary
    work_days = sc['work_days'] + sc['half_days']
    att_items = [
        (f'{work_days} Present',       0.00),
        (f'{sc["absent_days"]} Absent', 0.25),
        (f'{sc["sun_days"]} Sunday',    0.50),
        (f'{sc["total_ot_h"]:.1f}h OT', 0.75),
    ]
    for label, frac in att_items:
        txt(LX + frac * (RX - LX), y - 0.09 * inch, label, size=6.5)
    y -= 0.15 * inch
    hline(y, dashed=True)
    y -= 0.05 * inch

    # 4. Earnings rows
    R = 0.135 * inch
    earn_rows = [
        ('Basic Earned',                         f'Rs {sc["weekday_earned"]:,.0f}'),
        ('Sunday Bonus',                         f'Rs {sc["sun_base_pay"]+sc["sun_ot_pay"]+sc["iso_ot_pay"]:,.0f}'),
        (f'OT Pay  ({sc["total_ot_h"]:.1f}h)',   f'Rs {sc["weekday_ot_pay"]:,.0f}'),
        ('Advance Deduction',                    f'- Rs {sc["advance"]:,.0f}'),
    ]
    for label, val in earn_rows:
        ry = y - R
        txt(LX + 0.04 * inch, ry + R * 0.30, label, size=7)
        txt(RX - 0.04 * inch, ry + R * 0.30, val,   size=7, bold=True, align='right')
        y -= R

    hline(y, lw=0.8)
    y -= 0.05 * inch

    # 5. Net payable outlined box
    NET_H = 0.22 * inch
    outline_box(LX, y - NET_H, RX - LX, NET_H, lw=1.2)
    mid_y = y - NET_H / 2
    txt(LX + 0.08 * inch, mid_y - 0.022 * inch, 'NET PAYABLE AMOUNT', size=7, bold=True)
    txt(RX - 0.08 * inch, mid_y - 0.022 * inch, f'Rs {sc["net"]:,.0f}', size=10, bold=True, align='right')
    y -= NET_H + 0.06 * inch

    # 6. Signature + PAID stamp
    sig_x1 = RX - 1.10 * inch
    sig_y  = y - 0.20 * inch
    hline(sig_y, x1=sig_x1, x2=RX - 0.04 * inch, lw=0.5)
    txt((sig_x1 + RX - 0.04 * inch) / 2, sig_y - 0.09 * inch,
        'Authorised Signatory', size=6, align='center')
    if store.sig_path and os.path.exists(store.sig_path):
        try:
            c.drawImage(
                ImageReader(store.sig_path),
                sig_x1, sig_y + 0.02 * inch,
                width=(RX - 0.04 * inch - sig_x1), height=0.18 * inch,
                preserveAspectRatio=True, mask='auto',
            )
        except Exception:
            pass

    if info.get('paid', False):
        c.saveState()
        c.translate(LX + 0.38 * inch, sig_y - 0.02 * inch)
        c.rotate(8)
        c.setStrokeColor(BLACK)
        c.setFillColor(WHITE)
        c.setLineWidth(1.4)
        c.roundRect(-0.32 * inch, -0.11 * inch, 0.64 * inch, 0.22 * inch, 3, fill=1, stroke=1)
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(BLACK)
        c.drawCentredString(0, -0.034 * inch, 'PAID')
        c.restoreState()

    # 7. Footer
    hline(0.10 * inch, lw=0.4)
    c.setFont('Helvetica', 5.5)
    c.setFillColor(GRAY)
    c.drawCentredString(CX, 0.03 * inch,
                        f'Powered by {DEVELOPER_NAME}  \u2022  {DEVELOPER_PHONE}')

    c.save()
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
#  DIALOGS
# ─────────────────────────────────────────────────────────────────────────────
class SigSetupDialog(QDialog):
    """Let admin set the accountant signature image for PDF slips."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Signature Setup")
        self.setMinimumWidth(440)
        self.setStyleSheet(STYLE)
        lo = QVBoxLayout(self)
        lo.setContentsMargins(24, 24, 24, 24)
        lo.setSpacing(16)

        title = QLabel("Set Accountant Signature")
        title.setObjectName("heading")
        lo.addWidget(title)
        info = QLabel(
            "Upload a PNG/JPG image of the accountant's signature.\n"
            "It will appear on all salary slip PDFs."
        )
        info.setStyleSheet(f"color:{COLORS['muted']}; font-size:12px;")
        info.setWordWrap(True)
        lo.addWidget(info)

        self.path_lbl = QLabel(store.sig_path or "No signature set")
        self.path_lbl.setStyleSheet(
            f"background:{COLORS['surface']}; padding:8px 12px; border-radius:8px; "
            f"border:1px solid {COLORS['border']}; color:{COLORS['text']};"
        )
        self.path_lbl.setWordWrap(True)
        lo.addWidget(self.path_lbl)

        br = QHBoxLayout()
        browse_btn = QPushButton("Browse Image…")
        browse_btn.clicked.connect(self._browse)
        clear_btn  = QPushButton("Clear")
        clear_btn.setObjectName("danger")
        clear_btn.clicked.connect(self._clear)
        done_btn   = QPushButton("Done")
        done_btn.clicked.connect(self.accept)
        br.addWidget(browse_btn)
        br.addWidget(clear_btn)
        br.addStretch()
        br.addWidget(done_btn)
        lo.addLayout(br)

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Signature", "", "Images (*.png *.jpg *.jpeg *.bmp)"
        )
        if path:
            store.sig_path = path
            store.save()
            self.path_lbl.setText(path)

    def _clear(self):
        store.sig_path = ""
        store.save()
        self.path_lbl.setText("No signature set")


class EmpDialog(QDialog):
    """Edit employee name, salary, advance, paid status."""

    def __init__(self, eid: str, parent=None):
        super().__init__(parent)
        self.eid  = str(eid)
        info      = store.emp_info(eid)
        self.setWindowTitle(f"Employee — {eid}")
        self.setMinimumWidth(400)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(24, 24, 24, 24)
        lo.setSpacing(16)

        title = QLabel(f"Employee ID: {eid}")
        title.setObjectName("heading")
        lo.addWidget(title)

        form = QFormLayout()
        form.setSpacing(12)
        self.name_edit    = QLineEdit(info.get("name", str(eid)))
        self.salary_spin  = QDoubleSpinBox()
        self.salary_spin.setRange(0, 9_999_999)
        self.salary_spin.setDecimals(0)
        self.salary_spin.setSingleStep(500)
        self.salary_spin.setPrefix("Rs ")
        self.salary_spin.setValue(float(info.get("salary", 0)))
        self.advance_spin = QDoubleSpinBox()
        self.advance_spin.setRange(0, 9_999_999)
        self.advance_spin.setDecimals(0)
        self.advance_spin.setSingleStep(100)
        self.advance_spin.setPrefix("Rs ")
        self.advance_spin.setValue(float(info.get("advance", 0)))
        self.paid_chk     = QCheckBox("Salary paid this month")
        self.paid_chk.setChecked(info.get("paid", False))

        form.addRow("Name:",           self.name_edit)
        form.addRow("Monthly Salary:", self.salary_spin)
        form.addRow("Advance Taken:",  self.advance_spin)
        form.addRow("",                self.paid_chk)
        lo.addLayout(form)

        br = QHBoxLayout()
        save_btn   = QPushButton("Save")
        save_btn.clicked.connect(self._save)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("secondary")
        cancel_btn.clicked.connect(self.reject)
        br.addWidget(cancel_btn)
        br.addWidget(save_btn)
        lo.addLayout(br)

    def _save(self):
        store.set_emp(
            self.eid,
            self.name_edit.text().strip() or self.eid,
            self.salary_spin.value(),
            self.advance_spin.value(),
            self.paid_chk.isChecked(),
        )
        self.accept()


class HolidayDialog(QDialog):
    """Mark/unmark absent days as company holidays for one employee."""

    def __init__(self, eid: str, parent=None):
        super().__init__(parent)
        self.eid = str(eid)
        self.setWindowTitle(f"Mark Holidays — {store.emp_info(eid)['name']}")
        self.setMinimumSize(420, 500)
        self.setStyleSheet(STYLE)

        att = store.attendance.get(self.eid, {})
        lo  = QVBoxLayout(self)
        lo.setContentsMargins(20, 20, 20, 20)
        lo.setSpacing(14)

        lo.addWidget(QLabel("Check dates to mark as company holiday (no salary deduction):"))

        scroll_w  = QWidget()
        scroll_lo = QVBoxLayout(scroll_w)
        scroll_lo.setSpacing(4)
        scroll    = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(scroll_w)

        self.checks: Dict[str, QCheckBox] = {}
        # Show ALL dates (not just absents) so admin can mark any day
        for ds, rec in sorted(att.items()):
            dt  = datetime.strptime(ds, "%Y-%m-%d")
            lbl = f"{dt.strftime('%a, %d %b %Y')}  —  {STATUS_LABEL.get(rec['status'], rec['status'])}"
            chk = QCheckBox(lbl)
            chk.setChecked(rec.get("holiday", False))
            scroll_lo.addWidget(chk)
            self.checks[ds] = chk
        scroll_lo.addStretch()
        lo.addWidget(scroll)

        # Quick buttons: mark all Sundays / clear all
        qr = QHBoxLayout()
        mark_sun_btn = QPushButton("Mark All Sundays")
        mark_sun_btn.setObjectName("secondary")
        mark_sun_btn.clicked.connect(self._mark_all_sundays)
        clear_btn = QPushButton("Clear All")
        clear_btn.setObjectName("secondary")
        clear_btn.clicked.connect(self._clear_all)
        qr.addWidget(mark_sun_btn)
        qr.addWidget(clear_btn)
        lo.addLayout(qr)

        br = QHBoxLayout()
        save_btn   = QPushButton("Save")
        save_btn.clicked.connect(self._save)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("secondary")
        cancel_btn.clicked.connect(self.reject)
        br.addStretch()
        br.addWidget(cancel_btn)
        br.addWidget(save_btn)
        lo.addLayout(br)

    def _mark_all_sundays(self):
        for ds, chk in self.checks.items():
            if datetime.strptime(ds, "%Y-%m-%d").weekday() == 6:
                chk.setChecked(True)

    def _clear_all(self):
        for chk in self.checks.values():
            chk.setChecked(False)

    def _save(self):
        from attendance_engine import _weekday, ShiftState, apply_sunday_rules
        att = store.attendance.get(self.eid, {})
        for ds, chk in self.checks.items():
            if ds in att:
                att[ds]["holiday"] = chk.isChecked()
                if chk.isChecked():
                    att[ds]["status"] = DayStatus.HOLIDAY.value
                    att[ds]["note"]   = "Holiday marked by admin"
                else:
                    # Restore logical status based on day-of-week and existing shifts
                    dow = _weekday(ds)
                    payroll_shifts = [
                        s for s in att[ds].get("shifts", [])
                        if s.get("state") in (ShiftState.VALID.value, ShiftState.APPROVED.value)
                    ]
                    if payroll_shifts:
                        if dow == 6:
                            att[ds]["status"] = DayStatus.SUNDAY.value
                        elif dow == 5:
                            att[ds]["status"] = DayStatus.SATURDAY.value
                        else:
                            att[ds]["status"] = DayStatus.PRESENT.value
                    else:
                        if dow == 6:
                            att[ds]["status"] = DayStatus.SUNDAY_OFF.value
                        elif dow == 5:
                            att[ds]["status"] = DayStatus.SATURDAY_OFF.value
                        else:
                            att[ds]["status"] = DayStatus.ABSENT.value
                    att[ds]["note"] = "Holiday unmarked by admin"
        # Re-apply Sunday rules so adjacent-day changes (holiday → present) are reflected
        att = apply_sunday_rules(self.eid, att)
        store.attendance[self.eid] = att
        store.save()
        self.accept()


class ManualOverrideDialog(QDialog):
    """Admin manually overrides a specific day's attendance record."""

    def __init__(self, eid: str, date_str: str, parent=None):
        super().__init__(parent)
        self.eid      = str(eid)
        self.date_str = date_str
        self.setWindowTitle(f"Manual Override — {date_str}")
        self.setMinimumWidth(420)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(24, 24, 24, 24)
        lo.setSpacing(14)

        lo.addWidget(QLabel(f"Override attendance for {store.emp_info(eid)['name']} on {date_str}:"))

        form = QFormLayout()
        self.status_combo = QComboBox()
        self.status_combo.addItems([
            DayStatus.PRESENT.value, DayStatus.HALF_DAY.value,
            DayStatus.ABSENT.value, DayStatus.SATURDAY.value,
            DayStatus.SATURDAY_OFF.value, DayStatus.SUNDAY.value,
            DayStatus.SUNDAY_OFF.value, DayStatus.SUNDAY_ABSENT.value,
            DayStatus.HOLIDAY.value,
        ])
        rec = store.attendance.get(self.eid, {}).get(date_str, {})
        idx = self.status_combo.findText(rec.get("status", DayStatus.ABSENT.value))
        if idx >= 0:
            self.status_combo.setCurrentIndex(idx)

        self.worked_spin = QDoubleSpinBox()
        self.worked_spin.setRange(0, 24)
        self.worked_spin.setDecimals(2)
        self.worked_spin.setValue(float(rec.get("worked_h", 0)))

        self.ot_spin = QDoubleSpinBox()
        self.ot_spin.setRange(0, 16)
        self.ot_spin.setDecimals(2)
        self.ot_spin.setValue(float(rec.get("ot_h", 0)))

        self.note_edit = QLineEdit()
        self.note_edit.setPlaceholderText("Reason for override (optional)")

        form.addRow("Status:",        self.status_combo)
        form.addRow("Worked Hours:",  self.worked_spin)
        form.addRow("OT Hours:",      self.ot_spin)
        form.addRow("Note:",          self.note_edit)
        lo.addLayout(form)

        br = QHBoxLayout()
        apply_btn  = QPushButton("Apply Override")
        apply_btn.clicked.connect(self._apply)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("secondary")
        cancel_btn.clicked.connect(self.reject)
        br.addWidget(cancel_btn)
        br.addWidget(apply_btn)
        lo.addLayout(br)

    def _apply(self):
        ok = manual_override(
            store.attendance, self.eid, self.date_str,
            self.status_combo.currentText(),
            self.worked_spin.value(),
            self.ot_spin.value(),
            self.note_edit.text().strip(),
        )
        if ok:
            store.save()
            self.accept()
        else:
            QMessageBox.warning(self, "Error", f"Could not apply override for {self.date_str}")


class ManualPunchDialog(QDialog):
    """Admin adds a manual IN/OUT punch pair for an employee."""

    def __init__(self, eid: str, date_str: str, parent=None):
        super().__init__(parent)
        self.eid      = str(eid)
        self.date_str = date_str
        self.setWindowTitle(f"Add Manual Punch — {date_str}")
        self.setMinimumWidth(360)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(24, 24, 24, 24)
        lo.setSpacing(14)

        lo.addWidget(QLabel(f"Add IN/OUT punch for {store.emp_info(eid)['name']} on {date_str}:"))

        form = QFormLayout()
        self.in_edit  = QLineEdit("08:00")
        self.in_edit.setPlaceholderText("HH:MM")
        self.out_edit = QLineEdit("16:30")
        self.out_edit.setPlaceholderText("HH:MM")
        self.note_edit = QLineEdit("Manual entry by admin")
        form.addRow("IN time (HH:MM):",  self.in_edit)
        form.addRow("OUT time (HH:MM):", self.out_edit)
        form.addRow("Note:",             self.note_edit)
        lo.addLayout(form)

        br = QHBoxLayout()
        add_btn    = QPushButton("Add Punch")
        add_btn.clicked.connect(self._add)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("secondary")
        cancel_btn.clicked.connect(self.reject)
        br.addWidget(cancel_btn)
        br.addWidget(add_btn)
        lo.addLayout(br)

    def _add(self):
        try:
            ok = add_manual_punch(
                store.attendance, self.eid, self.date_str,
                self.in_edit.text().strip(),
                self.out_edit.text().strip(),
                self.note_edit.text().strip(),
            )
            if ok:
                store.save()
                self.accept()
            else:
                QMessageBox.warning(self, "Error", "Could not add punch. Check times.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


class SuspiciousShiftsDialog(QDialog):
    """Show all suspicious shifts across all employees. Admin can approve or review."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Suspicious Shifts — Admin Review")
        self.setMinimumSize(800, 500)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(20, 20, 20, 20)
        lo.setSpacing(14)

        lo.addWidget(QLabel("Shifts flagged for review (>16h). Approve to include in payroll:"))

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Emp ID", "Name", "Date", "IN", "OUT", "Raw Hours", "Action"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        lo.addWidget(self.table)

        self._populate()

        br = QHBoxLayout()
        close_btn = QPushButton("Close")
        close_btn.setObjectName("secondary")
        close_btn.clicked.connect(self.accept)
        br.addStretch()
        br.addWidget(close_btn)
        lo.addLayout(br)

    def _populate(self):
        entries = find_suspicious_shifts(store.attendance)
        self.table.setRowCount(len(entries))
        for row, entry in enumerate(entries):
            eid      = entry["eid"]
            date_str = entry["date_str"]
            shift    = entry["shift"]
            info     = store.emp_info(eid)
            vals = [
                eid, info.get("name", eid), date_str,
                shift.get("in", "-")[-8:-3] if shift.get("in") else "-",
                shift.get("out", "-")[-8:-3] if shift.get("out") else "-",
                f"{shift['raw_h']:.1f}h",
            ]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setForeground(QColor(COLORS["warning"]))
                self.table.setItem(row, col, item)
            # Approve button in last column
            btn = QPushButton("Approve")
            btn.setObjectName("success")
            btn.clicked.connect(lambda checked, e=eid, d=date_str: self._approve(e, d))
            self.table.setCellWidget(row, 6, btn)
            self.table.setRowHeight(row, 44)

    def _approve(self, eid: str, date_str: str):
        approve_suspicious_shift(store.attendance, eid, date_str)
        store.save()
        self._populate()
        QMessageBox.information(self, "Approved", f"Shift approved for {eid} on {date_str}.")


class AuditLogDialog(QDialog):
    """View the in-memory audit log."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Audit Log")
        self.setMinimumSize(900, 600)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(20, 20, 20, 20)
        lo.setSpacing(14)

        lo.addWidget(QLabel("Audit Log — all decisions made by the engine:"))

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet(
            f"background:{COLORS['surface']}; color:{COLORS['text']}; "
            f"font-family:monospace; font-size:11px; border-radius:8px;"
        )
        lo.addWidget(self.log_text)

        self._load_log()

        br = QHBoxLayout()
        export_btn = QPushButton("Export to JSON")
        export_btn.setObjectName("secondary")
        export_btn.clicked.connect(self._export)
        clear_btn  = QPushButton("Clear Log")
        clear_btn.setObjectName("danger")
        clear_btn.clicked.connect(self._clear)
        close_btn  = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        br.addWidget(export_btn)
        br.addWidget(clear_btn)
        br.addStretch()
        br.addWidget(close_btn)
        lo.addLayout(br)

    def _load_log(self):
        entries = get_audit_log()
        if not entries:
            self.log_text.setPlainText("(no audit entries yet)")
            return
        lines = []
        for e in entries:
            lvl = e.get("level", "info").upper()
            lines.append(f"[{e['ts']}] [{lvl:7s}] [{e['eid']:8s}] {e['msg']}")
        self.log_text.setPlainText("\n".join(lines))

    def _export(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Audit Log", "synix_audit.json", "JSON (*.json)"
        )
        if path:
            from attendance_engine import export_audit_log
            export_audit_log(path)
            QMessageBox.information(self, "Exported", f"Audit log saved to:\n{path}")

    def _clear(self):
        clear_audit_log()
        self.log_text.setPlainText("(log cleared)")


class TestRunnerDialog(QDialog):
    """Run all built-in unit tests and display results."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Unit Tests")
        self.setMinimumSize(700, 500)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(20, 20, 20, 20)
        lo.setSpacing(14)
        lo.addWidget(QLabel("Built-in Unit Test Results:"))

        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setStyleSheet(
            f"background:{COLORS['surface']}; color:{COLORS['text']}; "
            f"font-family:monospace; font-size:12px; border-radius:8px;"
        )
        lo.addWidget(self.result_text)

        self._run()

        br = QHBoxLayout()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        br.addStretch()
        br.addWidget(close_btn)
        lo.addLayout(br)

    def _run(self):
        att_results  = run_engine_tests()
        pay_results  = run_payroll_tests()
        all_results  = att_results + pay_results
        passed = sum(1 for r in all_results if r.startswith("[PASS]"))
        failed = sum(1 for r in all_results if r.startswith("[FAIL]"))

        lines = ["=== Attendance Engine Tests ==="]
        lines += att_results
        lines += ["", "=== Payroll Engine Tests ==="]
        lines += pay_results
        lines += [
            "",
            f"{'='*40}",
            f"TOTAL: {passed} PASSED, {failed} FAILED",
        ]
        self.result_text.setPlainText("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────────────
#  EMPLOYEE DETAIL DIALOG
# ─────────────────────────────────────────────────────────────────────────────
class EmpDetailDialog(QDialog):
    """
    Full-detail view for one employee:
    - Month attendance table (all days)
    - Salary breakdown
    - Admin actions: edit, holidays, override, manual punch, print slip
    """

    def __init__(self, eid: str, parent=None):
        super().__init__(parent)
        self.eid = str(eid)
        info     = store.emp_info(eid)
        self.setWindowTitle(f"Detail — {info['name']}")
        self.setMinimumSize(1040, 700)
        self.setStyleSheet(STYLE)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(20, 20, 20, 20)
        lo.setSpacing(14)

        # ── Header ──────────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        nl  = QLabel(info["name"])
        nl.setObjectName("heading")
        el  = QLabel(f"  ID #{eid}")
        el.setObjectName("subheading")
        hdr.addWidget(nl)
        hdr.addWidget(el)
        hdr.addStretch()
        edit_btn = QPushButton("Edit Info")
        edit_btn.setObjectName("secondary")
        edit_btn.clicked.connect(self._open_edit)
        hol_btn  = QPushButton("🗓 Mark Holidays")
        hol_btn.setObjectName("secondary")
        hol_btn.clicked.connect(self._open_holidays)
        hdr.addWidget(edit_btn)
        hdr.addWidget(hol_btn)
        lo.addLayout(hdr)

        # ── Stat cards ───────────────────────────────────────────────────────
        sc = self._get_sc()
        stats_row  = QHBoxLayout()
        self.c_work  = StatCard("Present Days",  sc["work_days"],        COLORS["success"])
        self.c_abs   = StatCard("Absent Days",   sc["absent_days"],      COLORS["danger"])
        self.c_sun   = StatCard("Sun Worked",    sc["sun_days"],         COLORS["warning"])
        self.c_oth   = StatCard("Total OT Hrs",  f'{sc["total_ot_h"]}h', COLORS["accent"])
        self.c_net   = StatCard("Net Salary",    f'Rs {sc["net"]:,.0f}', COLORS["success"])
        for card in [self.c_work, self.c_abs, self.c_sun, self.c_oth, self.c_net]:
            stats_row.addWidget(card)
        lo.addLayout(stats_row)

        # ── Salary summary ───────────────────────────────────────────────────
        self.sal_frame = QFrame()
        self.sal_frame.setObjectName("card")
        sfl = QHBoxLayout(self.sal_frame)
        sfl.setContentsMargins(16, 12, 16, 12)
        self.sal_lbl = QLabel()
        self.sal_lbl.setWordWrap(True)
        self.sal_lbl.setTextFormat(Qt.TextFormat.RichText)
        sfl.addWidget(self.sal_lbl)
        self._update_sal_label()
        lo.addWidget(self.sal_frame)

        # ── Attendance table ─────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Date", "Day", "IN", "OUT",
            "Worked Hrs", "OT Hrs", "Status", "Holiday", "Actions"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._fill_table()
        lo.addWidget(self.table)

        # ── Bottom buttons ───────────────────────────────────────────────────
        br = QHBoxLayout()
        self.paid_btn = QPushButton(
            "Mark as Paid ✓" if not info.get("paid") else "Mark as Unpaid"
        )
        self.paid_btn.setObjectName("success" if not info.get("paid") else "danger")
        self.paid_btn.clicked.connect(self._toggle_paid)

        slip_btn    = QPushButton("🖨  Print Salary Slip")
        slip_btn.setObjectName("secondary")
        slip_btn.clicked.connect(self._export_slip)

        thermal_btn = QPushButton("🔲  Thermal Print (TE244)")
        thermal_btn.setObjectName("secondary")
        thermal_btn.clicked.connect(self._export_thermal_slip)

        susp_btn  = QPushButton("⚠ Suspicious Shifts")
        susp_btn.setObjectName("warning")
        susp_btn.clicked.connect(self._show_suspicious)

        close_btn = QPushButton("Close")
        close_btn.setObjectName("secondary")
        close_btn.clicked.connect(self.accept)

        br.addStretch()
        br.addWidget(susp_btn)
        br.addWidget(thermal_btn)
        br.addWidget(slip_btn)
        br.addWidget(self.paid_btn)
        br.addWidget(close_btn)
        lo.addLayout(br)

    def _get_sc(self):
        return calc_salary(
            self.eid,
            store.emp_info(self.eid),
            store.attendance.get(self.eid, {}),
            store.month_label,
        )

    def _update_sal_label(self):
        sc   = self._get_sc()
        info = store.emp_info(self.eid)
        paid_icon = "✅" if info.get("paid") else "❌"
        note = f"  <span style='color:{COLORS['warning']}'>{sc['note']}</span>" if sc['note'] else ""
        text = (
            f"Per Day: <b>Rs {sc['per_day']:,.0f}</b>"
            f"  |  Per Hour: <b>Rs {sc['per_hour']:,.2f}</b>"
            f"  |  Weekday Earned: <b>Rs {sc['weekday_earned']:,.0f}</b>"
            + (f"  |  Sun Base ({sc['sun_days']}d): <b>Rs {sc['sun_base_pay']:,.0f}</b>"
               f"  |  Sun OT: <b>Rs {sc['sun_ot_pay']:,.0f}</b>" if sc['sun_days'] > 0 else "")
            + (f"  |  Sun Isolated ({sc['sun_iso_days']}d): <b>Rs {sc['iso_ot_pay']:,.0f}</b>"
               if sc['sun_iso_days'] > 0 else "")
            + (f"  |  Weekday OT ({sc['total_ot_h']}h): <b>Rs {sc['weekday_ot_pay']:,.0f}</b>"
               if sc['weekday_ot_pay'] > 0 else "")
            + f"  |  Absent Deduct: <b style='color:{COLORS['danger']}'>-Rs {sc['absent_deduction']:,.0f}</b>"
            + f"  |  Advance: <b style='color:{COLORS['danger']}'>-Rs {sc['advance']:,.0f}</b>"
            + f"  |  <b style='font-size:15px'>Gross: Rs {sc['gross']:,.0f}  →  Net: Rs {sc['net']:,.0f}</b>"
            + f"  |  Paid: {paid_icon}{note}"
        )
        self.sal_lbl.setText(text)
        self.sal_lbl.setStyleSheet(f"color:{COLORS['text']}; font-size:13px;")

    def _fill_table(self):
        att  = store.attendance.get(self.eid, {})
        rows = sorted(att.items())
        self.table.setRowCount(len(rows))

        for ri, (ds, rec) in enumerate(rows):
            dt     = datetime.strptime(ds, "%Y-%m-%d")
            status = rec.get("status", "")
            is_hol = rec.get("holiday", False)
            bg, fg = STATUS_BG_FG.get(status, (COLORS["bg"], COLORS["text"]))
            if is_hol:
                bg, fg = COLORS["holiday"], COLORS["success"]

            wh = rec.get("worked_h", 0)
            oh = rec.get("ot_h",     0)

            # Build suspicious badge for note
            susp_note = ""
            for sh in rec.get("shifts", []):
                if sh.get("state") == ShiftState.SUSPICIOUS.value:
                    susp_note = "⚠"
                    break

            cells = [
                ds,
                dt.strftime("%a"),
                rec.get("in",  "-"),
                rec.get("out", "-"),
                f"{wh:.2f}h" if wh > 0 else "-",
                f"{oh:.2f}h" if oh > 0 else "-",
                STATUS_LABEL.get(status, status) + (" " + susp_note if susp_note else ""),
                "🏖 Holiday" if is_hol else "",
            ]

            for col, val in enumerate(cells):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setForeground(QColor(fg))
                item.setBackground(QColor(bg))
                self.table.setItem(ri, col, item)

            # Action buttons column
            btn_widget = QWidget()
            btn_lo     = QHBoxLayout(btn_widget)
            btn_lo.setContentsMargins(2, 2, 2, 2)
            btn_lo.setSpacing(4)
            btn_widget.setStyleSheet(f"background:{bg};")

            override_btn = QPushButton("Override")
            override_btn.setObjectName("secondary")
            override_btn.setMaximumHeight(28)
            override_btn.clicked.connect(lambda ch, d=ds: self._override(d))

            punch_btn = QPushButton("+ Punch")
            punch_btn.setObjectName("secondary")
            punch_btn.setMaximumHeight(28)
            punch_btn.clicked.connect(lambda ch, d=ds: self._add_punch(d))

            btn_lo.addWidget(override_btn)
            btn_lo.addWidget(punch_btn)
            self.table.setCellWidget(ri, 8, btn_widget)
            self.table.setRowHeight(ri, 46)

    def _refresh(self):
        sc   = self._get_sc()
        info = store.emp_info(self.eid)
        self.c_work.update_value(sc["work_days"])
        self.c_abs.update_value(sc["absent_days"])
        self.c_sun.update_value(sc["sun_days"])
        self.c_oth.update_value(f'{sc["total_ot_h"]}h')
        self.c_net.update_value(f'Rs {sc["net"]:,.0f}')
        self._update_sal_label()
        self._fill_table()

    def _open_edit(self):
        if EmpDialog(self.eid, self).exec():
            self._refresh()

    def _open_holidays(self):
        if HolidayDialog(self.eid, self).exec():
            self._refresh()

    def _override(self, date_str: str):
        if ManualOverrideDialog(self.eid, date_str, self).exec():
            self._refresh()

    def _add_punch(self, date_str: str):
        if ManualPunchDialog(self.eid, date_str, self).exec():
            self._refresh()

    def _show_suspicious(self):
        SuspiciousShiftsDialog(self).exec()
        self._refresh()

    def _toggle_paid(self):
        info     = store.emp_info(self.eid)
        new_paid = not info.get("paid", False)
        store.set_emp(
            self.eid, info["name"], info["salary"], info["advance"], new_paid
        )
        if new_paid:
            self.paid_btn.setText("Mark as Unpaid")
            self.paid_btn.setObjectName("danger")
            # Ask user if they want to print the slip (don't force it)
            reply = QMessageBox.question(
                self, "Print Salary Slip?",
                f"Marked as Paid. Print salary slip for {info['name']}?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._export_slip()
        else:
            self.paid_btn.setText("Mark as Paid ✓")
            self.paid_btn.setObjectName("success")
        self.paid_btn.setStyleSheet(STYLE)
        self._update_sal_label()

    def _export_slip(self):
        info     = store.emp_info(self.eid)
        safe     = info.get("name", self.eid).replace(" ", "_")
        month    = (store.month_label or "slip").replace(" ", "_")
        default  = f"SalarySlip_{safe}_{month}.pdf"
        path, _  = QFileDialog.getSaveFileName(
            self, "Save Salary Slip PDF", default, "PDF files (*.pdf)"
        )
        if path:
            try:
                generate_salary_slip(self.eid, path)
                QMessageBox.information(self, "Done", f"Salary slip saved:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "PDF Error", str(e))

    def _export_thermal_slip(self):
        """Export TSC TE244 thermal-optimised salary slip (white bg, black ink only)."""
        info    = store.emp_info(self.eid)
        safe    = info.get("name", self.eid).replace(" ", "_")
        month   = (store.month_label or "slip").replace(" ", "_")
        default = f"ThermalSlip_{safe}_{month}.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Thermal Slip (TSC TE244)", default, "PDF files (*.pdf)"
        )
        if path:
            try:
                generate_thermal_slip(self.eid, path)
                QMessageBox.information(
                    self, "Done",
                    f"Thermal slip saved:\n{path}\n\n"
                    "TSC TE244 pe print karne ke liye:\n"
                    "1. TSC driver mein label size 4x3 inch set karein\n"
                    "2. Adobe Reader / SumatraPDF se print karein\n"
                    "3. 'Fit to page' option OFF rakhein"
                )
            except Exception as e:
                QMessageBox.critical(self, "PDF Error", str(e))


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN WINDOW
# ─────────────────────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    """
    Main application window.

    Layout
    ------
    TopBar: logo | month | signature | upload button
    Body:   Sidebar (summary stats) | Employee table (searchable, filterable)
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} — {COMPANY_NAME}")
        self.setMinimumSize(1150, 720)
        self.setStyleSheet(STYLE)
        self._worker: Optional[DatParserWorker] = None

        central = QWidget()
        self.setCentralWidget(central)
        root    = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Top bar ──────────────────────────────────────────────────────────
        topbar = QFrame()
        topbar.setStyleSheet(
            f"background:{COLORS['surface']}; border-bottom:1px solid {COLORS['border']};"
        )
        topbar.setFixedHeight(62)
        tb = QHBoxLayout(topbar)
        tb.setContentsMargins(20, 0, 20, 0)

        logo = QLabel("◈ Synix")
        logo.setStyleSheet(
            f"font-size:20px; font-weight:800; color:{COLORS['accent']}; letter-spacing:2px;"
        )
        sub = QLabel("Attendance & Payroll Manager")
        sub.setStyleSheet(f"font-size:11px; color:{COLORS['muted']}; margin-left:6px;")
        tb.addWidget(logo)
        tb.addWidget(sub)
        tb.addStretch()

        self.month_lbl = QLabel("No file loaded")
        self.month_lbl.setStyleSheet(f"color:{COLORS['muted']}; font-size:13px;")
        tb.addWidget(self.month_lbl)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setRange(0, 0)    # indeterminate spinner
        self.progress_bar.setMaximumWidth(100)
        self.progress_bar.setMaximumHeight(8)
        self.progress_bar.hide()
        tb.addWidget(self.progress_bar)

        sig_btn = QPushButton("✍  Signature")
        sig_btn.setObjectName("secondary")
        sig_btn.clicked.connect(lambda: SigSetupDialog(self).exec())
        tb.addWidget(sig_btn)

        print_btn = QPushButton("🖨  Print Payroll")
        print_btn.setObjectName("success")
        print_btn.clicked.connect(self._export_payroll_excel)
        tb.addWidget(print_btn)

        admin_btn = QPushButton("🔧 Admin")
        admin_btn.setObjectName("secondary")
        admin_btn.clicked.connect(self._show_admin_menu)
        tb.addWidget(admin_btn)

        up_btn = QPushButton("📂  Upload .dat")
        up_btn.clicked.connect(self._upload_file)
        tb.addWidget(up_btn)

        root.addWidget(topbar)

        # ── Body ─────────────────────────────────────────────────────────────
        body = QHBoxLayout()
        body.setContentsMargins(20, 16, 20, 16)
        body.setSpacing(16)
        root.addLayout(body)

        # Sidebar
        sidebar = QFrame()
        sidebar.setObjectName("card")
        sidebar.setMinimumWidth(200)
        sidebar.setMaximumWidth(260)
        sidebar.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        sb_lo = QVBoxLayout(sidebar)
        sb_lo.setContentsMargins(16, 16, 16, 16)
        sb_lo.setSpacing(12)
        sb_title = QLabel("Summary")
        sb_title.setObjectName("subheading")
        sb_lo.addWidget(sb_title)
        self.sb_total   = StatCard("Total Employees", "0",   COLORS["accent"])
        self.sb_present = StatCard("Avg Present",     "0d",  COLORS["success"])
        self.sb_absent  = StatCard("Avg Absent",      "0d",  COLORS["danger"])
        self.sb_ot      = StatCard("Total OT Hrs",    "0h",  COLORS["warning"])
        self.sb_paid    = StatCard("Paid",             "0",   COLORS["success"])
        self.sb_unpaid  = StatCard("Unpaid",           "0",   COLORS["danger"])
        for card in [self.sb_total, self.sb_present, self.sb_absent,
                     self.sb_ot, self.sb_paid, self.sb_unpaid]:
            sb_lo.addWidget(card)
        sb_lo.addStretch()
        body.addWidget(sidebar)

        # Main area
        ma = QVBoxLayout()
        ma.setSpacing(12)

        fr = QHBoxLayout()
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Search by name or employee ID...")
        self.search_box.textChanged.connect(self._filter_table)
        fr.addWidget(self.search_box)

        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["All Employees", "Unpaid Only", "Paid Only"])
        self.filter_combo.currentIndexChanged.connect(self._filter_table)
        fr.addWidget(self.filter_combo)
        ma.addLayout(fr)

        self.emp_table = QTableWidget()
        self.emp_table.setColumnCount(9)
        self.emp_table.setHorizontalHeaderLabels([
            "Emp ID", "Name", "Present", "Absent", "Sun Work",
            "Net Hrs", "OT Hrs", "Net Salary", "Status"
        ])
        self.emp_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.emp_table.verticalHeader().setVisible(False)
        self.emp_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.emp_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.emp_table.doubleClicked.connect(self._open_detail)
        ma.addWidget(self.emp_table)

        hint = QLabel(
            "Double-click a row → full attendance + salary detail  |  "
            "Admin → suspicious shifts, audit log, tests"
        )
        hint.setStyleSheet(f"color:{COLORS['muted']}; font-size:12px;")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ma.addWidget(hint)
        body.addLayout(ma)

        # Restore from disk if data exists
        if store.attendance:
            self.month_lbl.setText(store.month_label or "Loaded from cache")
            self._refresh_table()

    def _show_admin_menu(self):
        """Simple admin quick-access menu via message box (no QMenu needed)."""
        choices = ["Suspicious Shifts", "Audit Log", "Run Unit Tests", "Cancel"]
        msg = QMessageBox(self)
        msg.setWindowTitle("Admin Panel")
        msg.setText("Select admin action:")
        msg.setStyleSheet(STYLE)
        btns = [msg.addButton(c, QMessageBox.ButtonRole.AcceptRole) for c in choices]
        msg.exec()
        clicked = msg.clickedButton()
        if clicked == btns[0]:
            SuspiciousShiftsDialog(self).exec()
            self._refresh_table()
        elif clicked == btns[1]:
            AuditLogDialog(self).exec()
        elif clicked == btns[2]:
            TestRunnerDialog(self).exec()

    def _upload_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select .dat file", "", "DAT files (*.dat);;All files (*)"
        )
        if not path:
            return

        # Disable upload button and show progress during background parse
        existing_hols = store.get_all_holidays()
        self.progress_bar.show()
        self._worker = DatParserWorker(path, existing_hols)
        self._worker.finished.connect(self._on_parse_done)
        self._worker.error.connect(self._on_parse_error)
        self._worker.start()

    def _on_parse_done(self, att: dict, month_label: str):
        self.progress_bar.hide()
        if not att:
            QMessageBox.warning(
                self, "No Records",
                "No valid attendance records found.\n"
                "Expected format: EmployeeID  YYYY-MM-DD HH:MM:SS"
            )
            return
        store.attendance  = att
        store.month_label = month_label
        store.save()
        self.month_lbl.setText(f"📅 {month_label}")
        self._refresh_table()

        # Warn about missing salary employees
        missing = find_missing_salary_employees(store.employees, store.attendance)
        if missing:
            QMessageBox.information(
                self, "Salary Not Set",
                f"{len(missing)} employee(s) have no salary configured:\n"
                + ", ".join(missing[:10])
                + ("\n..." if len(missing) > 10 else "")
                + "\n\nDouble-click each employee to set their salary."
            )

    def _on_parse_error(self, msg: str):
        self.progress_bar.hide()
        QMessageBox.critical(self, "Parse Error", msg)

    def _refresh_table(self):
        self._filter_table()
        self._update_sidebar()

    def _update_sidebar(self):
        eids = store.all_eids()
        if not eids:
            return
        total    = len(eids)
        payrolls = calc_all_salaries(store.employees, store.attendance, store.month_label)
        avg_pres = sum(r["work_days"]    for r in payrolls.values()) / total
        avg_abs  = sum(r["absent_days"]  for r in payrolls.values()) / total
        tot_ot   = sum(r["total_ot_h"]   for r in payrolls.values())
        paid     = sum(1 for e in eids if store.emp_info(e).get("paid"))
        self.sb_total.update_value(total)
        self.sb_present.update_value(f"{avg_pres:.1f}d")
        self.sb_absent.update_value(f"{avg_abs:.1f}d")
        self.sb_ot.update_value(f"{tot_ot:.1f}h")
        self.sb_paid.update_value(paid)
        self.sb_unpaid.update_value(total - paid)

    def _filter_table(self):
        search = self.search_box.text().lower()
        ftype  = self.filter_combo.currentText()
        eids   = store.all_eids()
        rows   = []
        for eid in eids:
            info = store.emp_info(eid)
            if search and search not in info.get("name", "").lower() and search not in eid.lower():
                continue
            if ftype == "Unpaid Only" and info.get("paid"):
                continue
            if ftype == "Paid Only" and not info.get("paid"):
                continue
            rows.append(eid)

        self.emp_table.setRowCount(len(rows))
        for ri, eid in enumerate(rows):
            info    = store.emp_info(eid)
            att     = store.attendance.get(eid, {})
            sc      = calc_salary(eid, info, att, store.month_label)
            paid    = info.get("paid", False)

            vals = [
                eid,
                info.get("name", eid),
                str(sc["work_days"]),
                str(sc["absent_days"]),
                str(sc["sun_days"]),
                f'{sc["total_worked_h"]}h',
                f'{sc["total_ot_h"]}h',
                f'Rs {sc["net"]:,.0f}' if sc["salary"] > 0 else "Set salary →",
                "✅ Paid" if paid else "❌ Unpaid",
            ]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 3 and sc["absent_days"] > 5:
                    item.setForeground(QColor(COLORS["danger"]))
                elif col == 6 and sc["total_ot_h"] > 0:
                    item.setForeground(QColor(COLORS["warning"]))
                elif col == 7:
                    item.setForeground(
                        QColor(COLORS["success"] if sc["salary"] > 0 else COLORS["muted"])
                    )
                elif col == 8:
                    item.setForeground(
                        QColor(COLORS["success"] if paid else COLORS["danger"])
                    )
                else:
                    item.setForeground(QColor(COLORS["text"]))
                self.emp_table.setItem(ri, col, item)
            self.emp_table.setRowHeight(ri, 44)
            self.emp_table.item(ri, 0).setData(Qt.ItemDataRole.UserRole, eid)

    def _open_detail(self, index):
        item = self.emp_table.item(index.row(), 0)
        if not item:
            return
        eid = item.data(Qt.ItemDataRole.UserRole) or item.text()
        EmpDetailDialog(eid, self).exec()
        self._refresh_table()

    def _export_payroll_excel(self):
        """Export payroll summary for all employees as an attractive Excel file."""
        if not store.attendance:
            QMessageBox.warning(
                self, "No Data",
                "Koi attendance data nahi hai.\nPehle .dat file upload karein."
            )
            return
        month   = (store.month_label or "payroll").replace(" ", "_")
        default = f"Cosmo_Hydraulic_Payroll_{month}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Payroll Excel Report", default,
            "Excel files (*.xlsx)"
        )
        if not path:
            return
        try:
            generate_payroll_excel(path)
            QMessageBox.information(
                self, "✅ Export Successful",
                f"Payroll Excel report saved:\n{path}\n\n"
                "File mein sab employees ka data hai.\n"
                "Excel mein open karke print kar sakte hain."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )

    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setApplicationDisplayName(f"{APP_NAME} — {COMPANY_NAME}")
    app.setStyle("Fusion")

    pal = QPalette()
    pal.setColor(QPalette.ColorRole.Window,          QColor(COLORS["bg"]))
    pal.setColor(QPalette.ColorRole.WindowText,      QColor(COLORS["text"]))
    pal.setColor(QPalette.ColorRole.Base,            QColor(COLORS["surface"]))
    pal.setColor(QPalette.ColorRole.AlternateBase,   QColor(COLORS["card"]))
    pal.setColor(QPalette.ColorRole.Text,            QColor(COLORS["text"]))
    pal.setColor(QPalette.ColorRole.Button,          QColor(COLORS["surface"]))
    pal.setColor(QPalette.ColorRole.ButtonText,      QColor(COLORS["text"]))
    pal.setColor(QPalette.ColorRole.Highlight,       QColor(COLORS["accent"]))
    pal.setColor(QPalette.ColorRole.HighlightedText, QColor("#FFFFFF"))
    app.setPalette(pal)

    # ── License check (blocks until verified or user exits) ──────────────────
    dlg = LicenseCheckDialog()
    if dlg.exec() != QDialog.DialogCode.Accepted:
        sys.exit(0)   # user clicked Exit on blocked/no-internet dialog

    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()